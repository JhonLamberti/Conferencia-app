import re
import io
import unicodedata
from difflib import get_close_matches
from typing import Dict, List, Tuple, Optional, Set

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

EVENTOS_FIXOS_BASE = {
    "Faltas Por Hora": ["faltas por hora"],
    "Desconto DSR": ["desconto dsr"],
    "Faltas em Dia": ["faltas em dia"],
    "Repouso Remunerado": ["repouso remunerado"],
}

# -----------------------------
# Utilidades gerais
# -----------------------------

def normalizar_texto(texto: str) -> str:
    texto = unicodedata.normalize("NFD", str(texto or ""))
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto.lower().strip()


def normalizar_nome(nome: str) -> str:
    nome = normalizar_texto(nome)
    nome = re.sub(r"[^a-z0-9 ]", "", nome)
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome


def limpar_nome(nome: str) -> str:
    nome = re.sub(r"\s+", " ", nome or "").strip()
    # Corta qualquer informação que venha depois do nome na mesma linha.
    # Alguns PDFs extraem o cabeçalho do colaborador como:
    # "000044 IGOR ... 3.380,00 Função :..." ou com texto colado.
    nome = re.split(r"\s+\d{1,3}(?:\.\d{3})*,\d{2}\b", nome)[0].strip()
    nome = re.split(r"\bFun[cç][aã]o\b|\bAdmiss[aã]o\b|\bLivro\b|\bDep\s+IR\b", nome, flags=re.I)[0].strip()
    nome = re.sub(r"[:;,.]+$", "", nome).strip()
    return nome


def tempo_para_minutos(valor) -> int:
    if valor is None:
        return 0
    texto = str(valor).strip()
    if not texto:
        return 0
    m = re.search(r"(\d{1,4}):(\d{2})", texto)
    if not m:
        return 0
    return int(m.group(1)) * 60 + int(m.group(2))


def minutos_para_tempo(minutos: int) -> str:
    sinal = "-" if minutos < 0 else ""
    minutos = abs(int(minutos))
    return f"{sinal}{minutos // 60:02d}:{minutos % 60:02d}"


def formatar_dias(qtd) -> str:
    try:
        n = int(float(qtd))
    except Exception:
        n = 0
    return f"{n} dia" if n == 1 else f"{n} dias"


def decimal_ref_para_int(valor) -> int:
    """Converte referências do PDF como '001,00' ou '2,00' em quantidade de dias."""
    if valor is None or valor_vazio(valor):
        return 0
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return 0
    m = re.search(r"(\d{1,4})(?:[,.](\d{1,2}))?", texto)
    if not m:
        return 0
    try:
        return int(float(m.group(0).replace(".", "").replace(",", ".")))
    except Exception:
        try:
            return int(m.group(1))
        except Exception:
            return 0


def status_por_diferenca(diff_minutos: int) -> str:
    return "OK" if diff_minutos == 0 else "DIVERGENTE"


def valor_vazio(valor) -> bool:
    try:
        return pd.isna(valor)
    except Exception:
        return valor is None


def minutos_seguro(valor) -> int:
    """Converte valores do Excel/PDF em minutos sem quebrar quando vier vazio/NaN."""
    if valor is None or valor_vazio(valor):
        return 0
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return 0
    # quando vier algo como '01:35 Not.: 00:00', usa apenas o primeiro horário
    return tempo_para_minutos(texto)


def inteiro_seguro(valor) -> int:
    """Converte campos internos de minutos, tratando NaN como zero."""
    if valor is None or valor_vazio(valor):
        return 0
    try:
        return int(float(valor))
    except Exception:
        return minutos_seguro(valor)


def extrair_percentual(texto: str) -> Optional[str]:
    m = re.search(r"(\d{1,3})\s*%", str(texto or ""))
    return m.group(1) if m else None


def ordenar_percentuais(percentuais: Set[str]) -> List[str]:
    def chave(p):
        try:
            return int(p)
        except Exception:
            return 9999
    return sorted({str(p).replace('%', '').strip() for p in percentuais if str(p).strip()}, key=chave)

# -----------------------------
# Configuração manual de eventos
# -----------------------------

def montar_eventos_config(he_1_percent: str = "75", he_2_percent: str = "100", noturno_percent: str = "27") -> Dict[str, List[str]]:
    he_1_percent = str(he_1_percent).replace("%", "").strip()
    he_2_percent = str(he_2_percent).replace("%", "").strip()
    noturno_percent = str(noturno_percent).replace("%", "").strip()

    eventos = {
        f"Hora Extra {he_1_percent}%": [
            f"hora extra com {he_1_percent}%",
            f"hora extra {he_1_percent}%",
            f"he {he_1_percent}%",
            f"hora extra fator 1 ({he_1_percent}%)",
        ],
        f"Hora Extra {he_2_percent}%": [
            f"hora extra com {he_2_percent}%",
            f"hora extra {he_2_percent}%",
            f"he {he_2_percent}%",
            f"hora extra fator 2 ({he_2_percent}%)",
        ],
        f"Noturno {noturno_percent}%": [
            f"noturno {noturno_percent}%",
            f"adicional noturno {noturno_percent}%",
            "adicional noturno",
        ],
    }
    eventos.update(EVENTOS_FIXOS_BASE)
    return eventos


def montar_colunas_pdf(eventos: Dict[str, List[str]]) -> List[str]:
    colunas = ["Código", "Colaborador", "Página"]
    for evento in eventos:
        if evento == "Repouso Remunerado":
            colunas += [f"{evento} Valor"]
        else:
            colunas += [f"{evento} Ref", f"{evento} Valor"]
    return colunas

# -----------------------------
# Extração do PDF
# -----------------------------

def extrair_colaboradores_do_texto(texto: str, pagina: int) -> List[Dict]:
    linhas = [l.strip() for l in (texto or "").splitlines() if l.strip()]
    inicios: List[Tuple[int, str, str]] = []

    # Regra mais rígida para separar blocos por colaborador:
    # todo novo colaborador precisa iniciar uma nova linha com código de 6 dígitos.
    # Isso impede que eventos de um colaborador seguinte sejam anexados ao anterior.
    padrao_colaborador = re.compile(r"^\s*(\d{6})\s+(.+)$")

    for idx, linha in enumerate(linhas):
        m = padrao_colaborador.match(linha)
        if not m:
            continue

        codigo = m.group(1)
        resto = m.group(2).strip()

        # Evita falsos positivos em linhas que não são cabeçalho de colaborador.
        if not re.search(r"[A-ZÁÉÍÓÚÂÊÔÃÕÇ]{2,}", resto):
            continue

        nome = limpar_nome(resto)
        nome_norm = normalizar_texto(nome)

        # O nome deve ter pelo menos duas palavras.
        if len(nome_norm.split()) < 2:
            continue
        if any(x in nome_norm for x in ["empresa", "folha", "cnpj", "pagina", "codigo", "nome"]):
            continue

        # Remove duplicidade caso o mesmo cabeçalho seja lido duas vezes na página.
        if inicios and inicios[-1][0] == idx:
            continue
        inicios.append((idx, codigo, nome))

    colaboradores = []
    for pos, (idx, codigo, nome) in enumerate(inicios):
        fim = inicios[pos + 1][0] if pos + 1 < len(inicios) else len(linhas)
        bloco = "\n".join(linhas[idx:fim])
        colaboradores.append({"Código": codigo, "Colaborador": nome, "Página": pagina, "bloco": bloco})
    return colaboradores


def extrair_ref_valor_de_linha(linha: str) -> Tuple[str, str]:
    horas = re.findall(r"\b\d{1,4}:\d{2}\b", linha)
    valores = re.findall(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", linha)
    return (horas[-1] if horas else "", valores[-1] if valores else "")


def linha_dissidio_hora_extra(linha: str) -> bool:
    """Ignora rubricas de dissídio/reprocessamento ligadas a Hora Extra.

    Exemplo que NÃO deve entrar na conferência:
    '401 Dissídio 05/26 (1/1): Hora Extra 50% 0,02'

    A conferência deve considerar apenas a rubrica normal, como:
    '402 Hora Extra 100% 009:00 181,79'
    """
    norm = normalizar_texto(linha)
    return "dissidio" in norm and "hora extra" in norm


def extrair_ref_valor_evento(linha: str, evento: str) -> Tuple[str, str]:
    """Extrai referência e valor considerando o tipo de rubrica.

    Regras principais:
    - Hora extra, noturno e faltas por hora usam referência em horas (HH:MM).
    - Faltas em Dia e Desconto DSR usam referência em quantidade decimal (ex.: 001,00).
      Nessas linhas o primeiro decimal costuma ser o valor em R$ e o último decimal a referência.
    - Repouso Remunerado normalmente tem apenas valor em R$.
    """
    evento_norm = normalizar_texto(evento)
    horas = re.findall(r"\b\d{1,4}:\d{2}\b", linha)
    valores = re.findall(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", linha)

    if evento_norm in {"faltas em dia", "desconto dsr"}:
        # No PDF essas rubricas vêm no formato: código + descrição + referência + valor.
        # Ex.: "398 Faltas em Dia 001,00 67,49"
        # Portanto a referência é o primeiro decimal após a descrição e o valor é o último.
        ref = valores[0] if valores else ""
        valor = valores[-1] if len(valores) >= 2 else ""
        return ref, valor

    if evento_norm == "repouso remunerado":
        return "", valores[-1] if valores else ""

    return (horas[-1] if horas else "", valores[-1] if valores else "")


def extrair_evento_do_bloco(bloco: str, evento: str, aliases: List[str]) -> Tuple[str, str]:
    for linha in bloco.splitlines():
        linha_norm = normalizar_texto(linha)
        # Em eventos de Hora Extra, ignora linhas de dissídio para não capturar
        # valores residuais/reprocessados no lugar da HE normal do colaborador.
        if str(evento).lower().startswith("hora extra") and linha_dissidio_hora_extra(linha):
            continue
        # Também ignora dissídio para noturno e eventos fixos quando o nome da rubrica aparece no ajuste.
        if "dissidio" in linha_norm and any(normalizar_texto(alias) in linha_norm for alias in aliases):
            continue
        if any(normalizar_texto(alias) in linha_norm for alias in aliases):
            ref, valor = extrair_ref_valor_evento(linha, evento)
            if evento == "Repouso Remunerado":
                return "", valor
            return ref, valor
    return "", ""


def processar_pdf_manual(arquivo_pdf, eventos: Dict[str, List[str]], colunas_pdf: List[str]) -> pd.DataFrame:
    registros = []
    with pdfplumber.open(arquivo_pdf) as pdf:
        for numero_pagina, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for colab in extrair_colaboradores_do_texto(texto, numero_pagina):
                linha = {col: "" for col in colunas_pdf}
                linha["Código"] = colab["Código"]
                linha["Colaborador"] = colab["Colaborador"]
                linha["Página"] = colab["Página"]
                for evento, aliases in eventos.items():
                    ref, valor = extrair_evento_do_bloco(colab["bloco"], evento, aliases)
                    if evento == "Repouso Remunerado":
                        linha[f"{evento} Valor"] = valor
                    else:
                        linha[f"{evento} Ref"] = ref
                        linha[f"{evento} Valor"] = valor
                registros.append(linha)
    return pd.DataFrame(registros, columns=colunas_pdf)


def detectar_horas_extras_no_bloco(bloco: str) -> Dict[str, Dict[str, str]]:
    """Detecta qualquer linha de Hora Extra com percentual no bloco do colaborador."""
    encontrados: Dict[str, Dict[str, str]] = {}
    padrao_he = re.compile(r"hora\s+extra(?:\s+com)?(?:\s+fator\s*\d+)?\s*(?:\()?\s*(\d{1,3})\s*%", re.I)

    for linha in bloco.splitlines():
        # Ignora rubricas de dissídio ligadas a hora extra.
        # Elas aparecem como ajuste/reprocessamento e não devem ser comparadas
        # com o espelho de ponto.
        if linha_dissidio_hora_extra(linha):
            continue
        linha_norm = normalizar_texto(linha)
        m = padrao_he.search(linha_norm)
        if m:
            perc = m.group(1)
            if perc != "0":
                ref, valor = extrair_ref_valor_de_linha(linha)
                encontrados[perc] = {"Ref": ref, "Valor": valor}
    return encontrados


def detectar_noturnos_no_bloco(bloco: str) -> Dict[str, Dict[str, str]]:
    """Detecta Noturno/Adicional Noturno com percentual no PDF."""
    encontrados: Dict[str, Dict[str, str]] = {}
    padrao_noturno = re.compile(r"(?:adicional\s+)?noturno\s*(\d{1,3})\s*%", re.I)

    for linha in bloco.splitlines():
        linha_norm = normalizar_texto(linha)
        # Ignora rubricas de dissídio ligadas a noturno/adicional noturno,
        # assim como já fazemos com Hora Extra.
        if "dissidio" in linha_norm and "noturno" in linha_norm:
            continue
        m = padrao_noturno.search(linha_norm)
        if m:
            perc = m.group(1)
            ref, valor = extrair_ref_valor_de_linha(linha)
            encontrados[perc] = {"Ref": ref, "Valor": valor}
    return encontrados




def colunas_noturno_pdf(df_pdf: pd.DataFrame) -> List[str]:
    """Retorna todas as colunas de referência de Noturno do PDF, independente do percentual.

    Regra: qualquer evento lido como 'Noturno XX%' no PDF deve ser comparado
    contra a coluna 'Adicional noturno' da planilha.
    """
    return [c for c in df_pdf.columns if re.match(r"^Noturno\s+\d{1,3}%\s+Ref$", str(c))]


def total_noturno_pdf_min(row_pdf, df_pdf: pd.DataFrame) -> int:
    """Soma todos os Noturnos percentuais do PDF para o colaborador.

    Isso evita que o app pegue apenas o primeiro percentual existente no arquivo
    e deixe em branco quando o colaborador possui outro percentual.
    """
    total = 0
    for c in colunas_noturno_pdf(df_pdf):
        total += minutos_seguro(row_pdf.get(c, ""))
    return total

def processar_pdf_dinamico(arquivo_pdf) -> Tuple[pd.DataFrame, List[str]]:
    registros_base = []
    percentuais: Set[str] = set()
    percentuais_noturno: Set[str] = set()

    with pdfplumber.open(arquivo_pdf) as pdf:
        for numero_pagina, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for colab in extrair_colaboradores_do_texto(texto, numero_pagina):
                he_map = detectar_horas_extras_no_bloco(colab["bloco"])
                noturno_map = detectar_noturnos_no_bloco(colab["bloco"])
                percentuais.update(he_map.keys())
                percentuais_noturno.update(noturno_map.keys())

                # Eventos fixos continuam sendo extraídos junto.
                fixos = {}
                for evento, aliases in EVENTOS_FIXOS_BASE.items():
                    ref, valor = extrair_evento_do_bloco(colab["bloco"], evento, aliases)
                    if evento == "Repouso Remunerado":
                        fixos[f"{evento} Valor"] = valor
                    else:
                        fixos[f"{evento} Ref"] = ref
                        fixos[f"{evento} Valor"] = valor

                registros_base.append({
                    "Código": colab["Código"],
                    "Colaborador": colab["Colaborador"],
                    "Página": colab["Página"],
                    "_he_map": he_map,
                    "_noturno_map": noturno_map,
                    **fixos,
                })

    percentuais_ordenados = ordenar_percentuais(percentuais)
    colunas = ["Código", "Colaborador", "Página"]
    for p in percentuais_ordenados:
        colunas += [f"Hora Extra {p}% Ref", f"Hora Extra {p}% Valor"]
    percentuais_noturno_ordenados = ordenar_percentuais(percentuais_noturno)
    for p in percentuais_noturno_ordenados:
        colunas += [f"Noturno {p}% Ref", f"Noturno {p}% Valor"]
    for evento in EVENTOS_FIXOS_BASE:
        if evento == "Repouso Remunerado":
            colunas += [f"{evento} Valor"]
        else:
            colunas += [f"{evento} Ref", f"{evento} Valor"]

    linhas = []
    for reg in registros_base:
        linha = {col: "" for col in colunas}
        linha["Código"] = reg["Código"]
        linha["Colaborador"] = reg["Colaborador"]
        linha["Página"] = reg["Página"]
        for p in percentuais_ordenados:
            dados = reg.get("_he_map", {}).get(p, {})
            linha[f"Hora Extra {p}% Ref"] = dados.get("Ref", "")
            linha[f"Hora Extra {p}% Valor"] = dados.get("Valor", "")
        for p in percentuais_noturno_ordenados:
            dados = reg.get("_noturno_map", {}).get(p, {})
            linha[f"Noturno {p}% Ref"] = dados.get("Ref", "")
            linha[f"Noturno {p}% Valor"] = dados.get("Valor", "")
        for c, v in reg.items():
            if c in linha:
                linha[c] = v
        linhas.append(linha)

    return pd.DataFrame(linhas, columns=colunas), percentuais_ordenados

# -----------------------------
# Extração da planilha de ponto
# -----------------------------

def processar_planilha_ponto_manual(arquivo_excel, label_he_1: str, label_he_2: str, label_noturno: str, col_he_1: int = 10, col_he_2: int = 11, col_noturno: int = 12) -> pd.DataFrame:
    df_raw = pd.read_excel(arquivo_excel, header=None, dtype=str, engine="openpyxl").fillna("")
    registros = []
    idx_he_1 = max(col_he_1 - 1, 0)
    idx_he_2 = max(col_he_2 - 1, 0)
    idx_noturno = max(col_noturno - 1, 0)

    colaborador_atual = None
    faltas_dia_atual = 0
    faltas_hora_min_atual = 0
    idx_debito = 6  # Coluna G da planilha: horas em débito
    for _, row in df_raw.iterrows():
        col_a = str(row.iloc[0]).strip()
        col_b = str(row.iloc[1]).strip() if len(row) > 1 else ""
        col_a_norm = normalizar_texto(col_a)
        if col_a_norm == "colaborador" and col_b:
            colaborador_atual = col_b
            faltas_dia_atual = 0
            faltas_hora_min_atual = 0
            continue
        if colaborador_atual and col_a_norm not in {"totais", "data", "colaborador"}:
            debito_min = minutos_seguro(row.iloc[idx_debito] if len(row) > idx_debito else "")
            if debito_min >= 8 * 60:
                faltas_dia_atual += 1
            elif debito_min > 0:
                faltas_hora_min_atual += debito_min
        if colaborador_atual and col_a_norm == "totais":
            he_1 = row.iloc[idx_he_1] if len(row) > idx_he_1 else ""
            he_2 = row.iloc[idx_he_2] if len(row) > idx_he_2 else ""
            noturno = row.iloc[idx_noturno] if len(row) > idx_noturno else ""
            registros.append({
                "Colaborador Excel": colaborador_atual,
                "Chave Nome": normalizar_nome(colaborador_atual),
                f"{label_he_1} Excel": minutos_para_tempo(tempo_para_minutos(he_1)),
                f"{label_he_2} Excel": minutos_para_tempo(tempo_para_minutos(he_2)),
                f"{label_noturno} Excel": minutos_para_tempo(tempo_para_minutos(noturno)),
                f"{label_he_1} Excel Min": tempo_para_minutos(he_1),
                f"{label_he_2} Excel Min": tempo_para_minutos(he_2),
                f"{label_noturno} Excel Min": tempo_para_minutos(noturno),
                "Horas falta Excel": minutos_para_tempo(faltas_hora_min_atual),
                "Horas falta Excel Min": faltas_hora_min_atual,
                "Faltas em Dia Excel": formatar_dias(faltas_dia_atual),
                "Faltas em Dia Excel Qtd": faltas_dia_atual,
            })
            colaborador_atual = None
    return pd.DataFrame(registros)


def detectar_mapa_colunas_he_excel(row) -> Dict[str, int]:
    """Lê uma linha de cabeçalho e retorna {'75': 9, '100': 10, ...}."""
    mapa = {}
    for idx, cell in enumerate(row):
        texto = str(cell or "")
        norm = normalizar_texto(texto)
        if "hora" in norm and "extra" in norm:
            perc = extrair_percentual(texto)
            # cabeçalhos como "Horas extras fator 1 (0%)" existem em alguns blocos,
            # mas 0% não é uma HE válida para comparação.
            if perc and perc != "0":
                mapa[perc] = idx
    return mapa


def detectar_coluna_noturno_excel(row) -> Optional[int]:
    for idx, cell in enumerate(row):
        norm = normalizar_texto(cell)
        if "noturno" in norm:
            return idx
    return None


def processar_planilha_ponto_dinamica(arquivo_excel) -> Tuple[pd.DataFrame, List[str]]:
    """
    Detecta automaticamente colunas de hora extra pelo cabeçalho da planilha.
    Ex.: 'Horas extras fator 1 (75%)', 'Horas extras fator 2 (100%)', etc.
    Em cada bloco, pega os valores da linha TOTAIS.
    """
    df_raw = pd.read_excel(arquivo_excel, header=None, dtype=str, engine="openpyxl").fillna("")
    registros = []
    percentuais: Set[str] = set()

    colaborador_atual = None
    mapa_he_atual: Dict[str, int] = {}
    coluna_noturno_atual: Optional[int] = None
    faltas_dia_atual = 0
    faltas_hora_min_atual = 0
    idx_debito = 6  # Coluna G: débito diário do colaborador

    for _, row in df_raw.iterrows():
        col_a = str(row.iloc[0]).strip()
        col_b = str(row.iloc[1]).strip() if len(row) > 1 else ""
        col_a_norm = normalizar_texto(col_a)
        row_list = list(row)

        if col_a_norm == "colaborador" and col_b:
            colaborador_atual = col_b
            mapa_he_atual = {}
            coluna_noturno_atual = None
            faltas_dia_atual = 0
            faltas_hora_min_atual = 0
            continue

        # Cabeçalho do bloco, normalmente a linha com Data, Entradas, Horas extras...
        mapa_detectado = detectar_mapa_colunas_he_excel(row_list)
        if mapa_detectado:
            mapa_he_atual = mapa_detectado
            percentuais.update(mapa_detectado.keys())
            coluna_noturno_atual = detectar_coluna_noturno_excel(row_list)
            continue

        # Leitura da coluna G (Débito) dentro do bloco do colaborador.
        # Regra de auditoria: >= 08:00 conta 1 falta em dia; > 00:00 e < 08:00 soma como horas falta.
        if colaborador_atual and col_a_norm not in {"totais", "data", "colaborador"}:
            debito_min = minutos_seguro(row.iloc[idx_debito] if len(row) > idx_debito else "")
            if debito_min >= 8 * 60:
                faltas_dia_atual += 1
            elif debito_min > 0:
                faltas_hora_min_atual += debito_min

        if colaborador_atual and col_a_norm == "totais":
            reg = {
                "Colaborador Excel": colaborador_atual,
                "Chave Nome": normalizar_nome(colaborador_atual),
                "Horas falta Excel": minutos_para_tempo(faltas_hora_min_atual),
                "Horas falta Excel Min": faltas_hora_min_atual,
                "Faltas em Dia Excel": formatar_dias(faltas_dia_atual),
                "Faltas em Dia Excel Qtd": faltas_dia_atual,
            }
            percentuais_do_colaborador = []
            for p, idx_col in mapa_he_atual.items():
                valor = row.iloc[idx_col] if len(row) > idx_col else ""
                minutos = minutos_seguro(valor)
                reg[f"Hora Extra {p}% Excel"] = minutos_para_tempo(minutos)
                reg[f"Hora Extra {p}% Excel Min"] = minutos
                percentuais_do_colaborador.append(p)
            reg["Percentuais HE Excel"] = ",".join(ordenar_percentuais(set(percentuais_do_colaborador)))
            if coluna_noturno_atual is not None:
                valor_noturno = row.iloc[coluna_noturno_atual] if len(row) > coluna_noturno_atual else ""
                minutos_noturno = minutos_seguro(valor_noturno)
                reg["Adicional Noturno Excel"] = minutos_para_tempo(minutos_noturno)
                reg["Adicional Noturno Excel Min"] = minutos_noturno
            registros.append(reg)
            colaborador_atual = None

    percentuais_ordenados = ordenar_percentuais(percentuais)
    colunas = ["Colaborador Excel", "Chave Nome", "Percentuais HE Excel"]
    for p in percentuais_ordenados:
        colunas += [f"Hora Extra {p}% Excel", f"Hora Extra {p}% Excel Min"]
    colunas += ["Adicional Noturno Excel", "Adicional Noturno Excel Min", "Horas falta Excel", "Horas falta Excel Min", "Faltas em Dia Excel", "Faltas em Dia Excel Qtd"]

    df = pd.DataFrame(registros)
    for c in colunas:
        if c not in df.columns:
            if c.endswith(" Min") or c.endswith(" Qtd"):
                df[c] = 0
            elif c == "Percentuais HE Excel":
                df[c] = ""
            elif "Dia" in c:
                df[c] = "0 dias"
            else:
                df[c] = "00:00"
    # evita erro "cannot convert float NaN to integer" nas comparações
    for c in df.columns:
        if c.endswith(" Min") or c.endswith(" Qtd"):
            df[c] = df[c].apply(inteiro_seguro)
        else:
            df[c] = df[c].fillna("")
    return df[colunas] if not df.empty else pd.DataFrame(columns=colunas), percentuais_ordenados

# -----------------------------
# Comparações
# -----------------------------

def encontrar_colaborador_excel(nome_pdf: str, mapa_excel: Dict[str, dict]) -> Optional[dict]:
    chave = normalizar_nome(nome_pdf)
    if chave in mapa_excel:
        return mapa_excel[chave]
    chaves = list(mapa_excel.keys())
    similares = get_close_matches(chave, chaves, n=1, cutoff=0.92)
    if similares:
        return mapa_excel[similares[0]]
    return None


def comparar_manual(df_pdf: pd.DataFrame, df_excel: pd.DataFrame, label_he_1: str, label_he_2: str, label_noturno: str) -> pd.DataFrame:
    colunas = [
        "Colaborador PDF", "Colaborador Excel", "Código", "Página",
        f"{label_he_1} PDF", f"{label_he_1} Excel", f"Diferença {label_he_1}", f"Status {label_he_1}",
        f"{label_he_2} PDF", f"{label_he_2} Excel", f"Diferença {label_he_2}", f"Status {label_he_2}",
        f"{label_noturno} PDF", f"{label_noturno} Excel", f"Diferença {label_noturno}", f"Status {label_noturno}",
        "Horas falta PDF", "Horas falta Excel", "Diferença Horas falta", "Status Horas falta",
        "Faltas em Dia PDF", "Faltas em Dia Excel", "Diferença Faltas em Dia", "Status Faltas em Dia",
        "Status Geral",
    ]
    if df_excel.empty:
        return pd.DataFrame(columns=colunas)

    mapa_excel = {row["Chave Nome"]: row for _, row in df_excel.iterrows()}
    registros = []
    for _, row_pdf in df_pdf.iterrows():
        nome_pdf = row_pdf.get("Colaborador", "")
        match = encontrar_colaborador_excel(nome_pdf, mapa_excel)
        pdf_he_1_min = tempo_para_minutos(row_pdf.get(f"{label_he_1} Ref", ""))
        pdf_he_2_min = tempo_para_minutos(row_pdf.get(f"{label_he_2} Ref", ""))
        pdf_noturno_min = tempo_para_minutos(row_pdf.get(f"{label_noturno} Ref", ""))

        if match is None:
            registros.append({
                "Colaborador PDF": nome_pdf, "Colaborador Excel": "NÃO ENCONTRADO", "Código": row_pdf.get("Código", ""), "Página": row_pdf.get("Página", ""),
                f"{label_he_1} PDF": minutos_para_tempo(pdf_he_1_min), f"{label_he_1} Excel": "", f"Diferença {label_he_1}": "", f"Status {label_he_1}": "NÃO ENCONTRADO",
                f"{label_he_2} PDF": minutos_para_tempo(pdf_he_2_min), f"{label_he_2} Excel": "", f"Diferença {label_he_2}": "", f"Status {label_he_2}": "NÃO ENCONTRADO",
                f"{label_noturno} PDF": minutos_para_tempo(pdf_noturno_min), f"{label_noturno} Excel": "", f"Diferença {label_noturno}": "", f"Status {label_noturno}": "NÃO ENCONTRADO",
                "Horas falta PDF": minutos_para_tempo(minutos_seguro(row_pdf.get("Faltas Por Hora Ref", ""))), "Horas falta Excel": "", "Diferença Horas falta": "", "Status Horas falta": "NÃO ENCONTRADO",
                "Faltas em Dia PDF": formatar_dias(decimal_ref_para_int(row_pdf.get("Faltas em Dia Ref", ""))), "Faltas em Dia Excel": "", "Diferença Faltas em Dia": "", "Status Faltas em Dia": "NÃO ENCONTRADO",
                "Status Geral": "COLABORADOR NÃO ENCONTRADO NO EXCEL",
            })
            continue

        vals = []
        for label, pdf_min in [(label_he_1, pdf_he_1_min), (label_he_2, pdf_he_2_min), (label_noturno, pdf_noturno_min)]:
            excel_min = inteiro_seguro(match.get(f"{label} Excel Min", 0))
            diff = pdf_min - excel_min
            vals += [minutos_para_tempo(pdf_min), minutos_para_tempo(excel_min), minutos_para_tempo(diff), status_por_diferenca(diff)]
        pdf_fh_min = minutos_seguro(row_pdf.get("Faltas Por Hora Ref", ""))
        excel_fh_min = inteiro_seguro(match.get("Horas falta Excel Min", 0))
        diff_fh = pdf_fh_min - excel_fh_min
        status_fh = status_por_diferenca(diff_fh)

        pdf_fd_qtd = decimal_ref_para_int(row_pdf.get("Faltas em Dia Ref", ""))
        excel_fd_qtd = inteiro_seguro(match.get("Faltas em Dia Excel Qtd", 0))
        diff_fd = pdf_fd_qtd - excel_fd_qtd
        status_fd = status_por_diferenca(diff_fd)

        status_geral = "OK" if vals[3] == vals[7] == vals[11] == status_fh == status_fd == "OK" else "DIVERGENTE"
        registros.append({
            "Colaborador PDF": nome_pdf, "Colaborador Excel": match.get("Colaborador Excel", ""), "Código": row_pdf.get("Código", ""), "Página": row_pdf.get("Página", ""),
            f"{label_he_1} PDF": vals[0], f"{label_he_1} Excel": vals[1], f"Diferença {label_he_1}": vals[2], f"Status {label_he_1}": vals[3],
            f"{label_he_2} PDF": vals[4], f"{label_he_2} Excel": vals[5], f"Diferença {label_he_2}": vals[6], f"Status {label_he_2}": vals[7],
            f"{label_noturno} PDF": vals[8], f"{label_noturno} Excel": vals[9], f"Diferença {label_noturno}": vals[10], f"Status {label_noturno}": vals[11],
            "Horas falta PDF": minutos_para_tempo(pdf_fh_min), "Horas falta Excel": minutos_para_tempo(excel_fh_min), "Diferença Horas falta": minutos_para_tempo(diff_fh), "Status Horas falta": status_fh,
            "Faltas em Dia PDF": formatar_dias(pdf_fd_qtd), "Faltas em Dia Excel": formatar_dias(excel_fd_qtd), "Diferença Faltas em Dia": formatar_dias(diff_fd), "Status Faltas em Dia": status_fd,
            "Status Geral": status_geral,
        })
    return pd.DataFrame(registros, columns=colunas)


def comparar_dinamico(df_pdf: pd.DataFrame, df_excel: pd.DataFrame, percentuais: List[str], comparar_noturno: bool = True) -> pd.DataFrame:
    base_cols = ["Colaborador PDF", "Colaborador Excel", "Código", "Página"]
    cols = base_cols[:]
    for p in percentuais:
        label = f"Hora Extra {p}%"
        cols += [f"{label} PDF", f"{label} Excel", f"Diferença {label}", f"Status {label}"]
    if comparar_noturno:
        cols += ["Adicional Noturno PDF", "Adicional Noturno Excel", "Diferença Adicional Noturno", "Status Adicional Noturno"]
    cols += ["Horas falta PDF", "Horas falta Excel", "Diferença Horas falta", "Status Horas falta"]
    cols += ["Faltas em Dia PDF", "Faltas em Dia Excel", "Diferença Faltas em Dia", "Status Faltas em Dia"]
    cols += ["Status Geral"]

    if df_excel.empty:
        return pd.DataFrame(columns=cols)

    mapa_excel = {row["Chave Nome"]: row for _, row in df_excel.iterrows()}
    registros = []
    for _, row_pdf in df_pdf.iterrows():
        nome_pdf = row_pdf.get("Colaborador", "")
        match = encontrar_colaborador_excel(nome_pdf, mapa_excel)
        reg = {"Colaborador PDF": nome_pdf, "Colaborador Excel": "NÃO ENCONTRADO" if match is None else match.get("Colaborador Excel", ""), "Código": row_pdf.get("Código", ""), "Página": row_pdf.get("Página", "")}
        status_list = []

        if match is None:
            for p in percentuais:
                label = f"Hora Extra {p}%"
                pdf_min = tempo_para_minutos(row_pdf.get(f"{label} Ref", ""))
                reg[f"{label} PDF"] = minutos_para_tempo(pdf_min)
                reg[f"{label} Excel"] = ""
                reg[f"Diferença {label}"] = ""
                reg[f"Status {label}"] = "NÃO ENCONTRADO"
            if comparar_noturno:
                reg["Adicional Noturno PDF"] = minutos_para_tempo(total_noturno_pdf_min(row_pdf, df_pdf))
                reg["Adicional Noturno Excel"] = ""
                reg["Diferença Adicional Noturno"] = ""
                reg["Status Adicional Noturno"] = "NÃO ENCONTRADO"
            reg["Horas falta PDF"] = minutos_para_tempo(minutos_seguro(row_pdf.get("Faltas Por Hora Ref", "")))
            reg["Horas falta Excel"] = ""
            reg["Diferença Horas falta"] = ""
            reg["Status Horas falta"] = "NÃO ENCONTRADO"
            reg["Faltas em Dia PDF"] = formatar_dias(decimal_ref_para_int(row_pdf.get("Faltas em Dia Ref", "")))
            reg["Faltas em Dia Excel"] = ""
            reg["Diferença Faltas em Dia"] = ""
            reg["Status Faltas em Dia"] = "NÃO ENCONTRADO"
            reg["Status Geral"] = "COLABORADOR NÃO ENCONTRADO NO EXCEL"
            registros.append(reg)
            continue

        percentuais_excel_colab = set(str(match.get("Percentuais HE Excel", "") or "").split(","))
        percentuais_excel_colab.discard("")
        for p in percentuais:
            label = f"Hora Extra {p}%"
            pdf_min = minutos_seguro(row_pdf.get(f"{label} Ref", ""))
            excel_tem_coluna = p in percentuais_excel_colab
            excel_min = inteiro_seguro(match.get(f"{label} Excel Min", 0)) if excel_tem_coluna else 0
            diff = pdf_min - excel_min
            if not excel_tem_coluna and pdf_min != 0:
                status = "SEM COLUNA NO EXCEL"
            else:
                status = status_por_diferenca(diff)
            status_list.append(status)
            reg[f"{label} PDF"] = minutos_para_tempo(pdf_min)
            reg[f"{label} Excel"] = minutos_para_tempo(excel_min) if excel_tem_coluna else "SEM COLUNA"
            reg[f"Diferença {label}"] = minutos_para_tempo(diff) if excel_tem_coluna or pdf_min != 0 else ""
            reg[f"Status {label}"] = status

        if comparar_noturno:
            # Qualquer "Noturno XX%" identificado no PDF deve ser associado
            # ao "Adicional Noturno" da planilha. Como alguns contratos podem
            # ter percentuais diferentes, somamos todos os Noturnos percentuais
            # encontrados no bloco daquele colaborador.
            pdf_noturno_min = total_noturno_pdf_min(row_pdf, df_pdf)
            excel_noturno_min = inteiro_seguro(match.get("Adicional Noturno Excel Min", 0))
            diff = pdf_noturno_min - excel_noturno_min
            status = status_por_diferenca(diff)
            status_list.append(status)
            reg["Adicional Noturno PDF"] = minutos_para_tempo(pdf_noturno_min)
            reg["Adicional Noturno Excel"] = minutos_para_tempo(excel_noturno_min)
            reg["Diferença Adicional Noturno"] = minutos_para_tempo(diff)
            reg["Status Adicional Noturno"] = status

        # Comparação das faltas calculadas pela coluna G (Débito) da planilha.
        pdf_fh_min = minutos_seguro(row_pdf.get("Faltas Por Hora Ref", ""))
        excel_fh_min = inteiro_seguro(match.get("Horas falta Excel Min", 0))
        diff_fh = pdf_fh_min - excel_fh_min
        status_fh = status_por_diferenca(diff_fh)
        status_list.append(status_fh)
        reg["Horas falta PDF"] = minutos_para_tempo(pdf_fh_min)
        reg["Horas falta Excel"] = minutos_para_tempo(excel_fh_min)
        reg["Diferença Horas falta"] = minutos_para_tempo(diff_fh)
        reg["Status Horas falta"] = status_fh

        pdf_fd_qtd = decimal_ref_para_int(row_pdf.get("Faltas em Dia Ref", ""))
        excel_fd_qtd = inteiro_seguro(match.get("Faltas em Dia Excel Qtd", 0))
        diff_fd = pdf_fd_qtd - excel_fd_qtd
        status_fd = status_por_diferenca(diff_fd)
        status_list.append(status_fd)
        reg["Faltas em Dia PDF"] = formatar_dias(pdf_fd_qtd)
        reg["Faltas em Dia Excel"] = formatar_dias(excel_fd_qtd)
        reg["Diferença Faltas em Dia"] = formatar_dias(diff_fd)
        reg["Status Faltas em Dia"] = status_fd

        reg["Status Geral"] = "OK" if all(s == "OK" for s in status_list) else "DIVERGENTE"
        registros.append(reg)

    return pd.DataFrame(registros, columns=cols)

# -----------------------------
# Exportação
# -----------------------------

def ajustar_larguras(ws):
    """Ajusta as larguras de forma controlada para o Excel não ficar visualmente quebrado."""
    larguras_padrao = {
        "Colaborador": 34,
        "Evento": 24,
        "PDF": 14,
        "Planilha": 16,
        "Diferença": 16,
        "Status": 28,
        "Motivo": 72,
        "Código": 12,
        "Página": 10,
        "Indicador": 42,
        "Quantidade": 16,
    }

    for col in ws.columns:
        letra = col[0].column_letter
        cabecalho = str(ws.cell(row=1, column=col[0].column).value or "").strip()
        if cabecalho in larguras_padrao:
            largura = larguras_padrao[cabecalho]
        else:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            largura = min(max(max_len + 2, 12), 34)
        ws.column_dimensions[letra].width = largura


def adicionar_tabela_excel(ws):
    """Cria tabela com filtro e estilo profissional, sem quebrar se o nome da aba tiver acentos."""
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    nome_base = re.sub(r"[^A-Za-z0-9_]", "", ws.title) or "Tabela"
    nome_tabela = f"tbl_{nome_base[:20]}"
    try:
        tab = Table(displayName=nome_tabela, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        # Se já houver tabela ou algum detalhe impedir a criação, mantém filtros normais.
        ws.auto_filter.ref = ref


def aplicar_estilo_planilha(wb):
    """Formata o Excel exportado para ficar limpo, filtrável e fácil de auditar."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill("solid", fgColor="FCE4D6")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    blue_fill = PatternFill("solid", fgColor="DDEBF7")
    gray_fill = PatternFill("solid", fgColor="E7E6E6")
    purple_fill = PatternFill("solid", fgColor="E4DFEC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

        headers = [str(c.value or "").strip() for c in ws[1]] if ws.max_row >= 1 else []
        header_map = {h: i + 1 for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 22
            status_text = str(ws.cell(row=row[0].row, column=header_map.get("Status", 1)).value or "")
            evento_text = str(ws.cell(row=row[0].row, column=header_map.get("Evento", 1)).value or "")
            norm_status = normalizar_texto(status_text)
            norm_evento = normalizar_texto(evento_text)

            fill = None
            if ws.title == "Divergências":
                if "sem coluna" in norm_status or "nao encontrado" in norm_status:
                    fill = yellow_fill
                elif norm_evento == "horas falta":
                    fill = purple_fill
                else:
                    fill = red_fill
            elif ws.title == "Horas falta":
                fill = purple_fill
            elif "ok" in norm_status and "divergente" not in norm_status:
                fill = green_fill

            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                if fill:
                    cell.fill = fill

        # Alinhamentos específicos para leitura rápida
        for col_name in ["PDF", "Planilha", "Diferença", "Quantidade", "Código", "Página"]:
            idx = header_map.get(col_name)
            if idx:
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=ws.max_row):
                    for c in cell:
                        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        idx_motivo = header_map.get("Motivo")
        if idx_motivo:
            for c in ws.iter_cols(min_col=idx_motivo, max_col=idx_motivo, min_row=2, max_row=ws.max_row):
                for cell in c:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ajustar_larguras(ws)
        adicionar_tabela_excel(ws)

    if "Resumo" in wb.sheetnames:
        ws = wb["Resumo"]
        ws.sheet_properties.tabColor = "70AD47"
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = blue_fill if cell.column == 1 else gray_fill
                cell.font = Font(bold=True if cell.column == 1 else False)
                cell.border = border
                cell.alignment = Alignment(vertical="center")

    if "Divergências" in wb.sheetnames:
        wb["Divergências"].sheet_properties.tabColor = "C00000"
    if "Horas falta" in wb.sheetnames:
        wb["Horas falta"].sheet_properties.tabColor = "7030A0"
    if "Completo" in wb.sheetnames:
        wb["Completo"].sheet_properties.tabColor = "A5A5A5"


def status_eh_problema(status: str) -> bool:
    texto = normalizar_texto(status)
    if not texto or texto == "ok":
        return False
    return True


def montar_divergencias_limpas(df_comparacao: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Transforma a comparação completa em uma visão operacional: somente problemas, uma linha por evento."""
    colunas = ["Colaborador", "Evento", "PDF", "Planilha", "Diferença", "Status", "Motivo", "Código", "Página"]
    if df_comparacao is None or df_comparacao.empty:
        return pd.DataFrame(columns=colunas)

    registros = []
    for _, row in df_comparacao.iterrows():
        nome_pdf = row.get("Colaborador PDF", "")
        nome_excel = row.get("Colaborador Excel", "")
        codigo = row.get("Código", "")
        pagina = row.get("Página", "")
        status_geral = str(row.get("Status Geral", ""))

        # Caso especial: colaborador nem apareceu na planilha.
        if normalizar_texto(status_geral).startswith("colaborador nao encontrado"):
            registros.append({
                "Colaborador": nome_pdf,
                "Evento": "Colaborador",
                "PDF": "Encontrado",
                "Planilha": "Não encontrado",
                "Diferença": "",
                "Status": "COLABORADOR NÃO ENCONTRADO",
                "Motivo": "Colaborador existe no PDF, mas não foi localizado na planilha de ponto.",
                "Código": codigo,
                "Página": pagina,
            })
            continue

        for coluna_status in [c for c in df_comparacao.columns if c.startswith("Status ") and c != "Status Geral"]:
            status = str(row.get(coluna_status, ""))
            if not status_eh_problema(status):
                continue

            evento = coluna_status.replace("Status ", "", 1)
            valor_pdf = row.get(f"{evento} PDF", "")
            valor_excel = row.get(f"{evento} Excel", "")
            diferenca = row.get(f"Diferença {evento}", "")

            status_norm = normalizar_texto(status)
            if "sem coluna" in status_norm:
                motivo = "Evento existe no PDF, mas não foi encontrada coluna correspondente na planilha para este colaborador."
            elif "nao encontrado" in status_norm:
                motivo = "Colaborador ou evento não localizado na planilha."
            elif normalizar_texto(evento) == "horas falta":
                motivo = "Comparação entre Faltas Por Hora do PDF e Horas falta calculadas pela coluna G da planilha. Na coluna G, valores > 00:00 e < 08:00 são somados como Horas falta."
            else:
                motivo = "Valor do PDF diferente do valor da planilha."

            registros.append({
                "Colaborador": nome_pdf or nome_excel,
                "Evento": evento,
                "PDF": valor_pdf,
                "Planilha": valor_excel,
                "Diferença": diferenca,
                "Status": status,
                "Motivo": motivo,
                "Código": codigo,
                "Página": pagina,
            })

    return pd.DataFrame(registros, columns=colunas)




def filtrar_horas_falta(df_divergencias: pd.DataFrame) -> pd.DataFrame:
    """Retorna apenas divergências de Horas falta para destaque na tela/Excel."""
    if df_divergencias is None or df_divergencias.empty or "Evento" not in df_divergencias.columns:
        return pd.DataFrame(columns=df_divergencias.columns if df_divergencias is not None else [])
    return df_divergencias[df_divergencias["Evento"].astype(str).str.lower().str.strip() == "horas falta"].copy()

def montar_resumo(df_pdf: pd.DataFrame, df_excel: Optional[pd.DataFrame], df_comparacao: Optional[pd.DataFrame], df_divergencias: pd.DataFrame) -> pd.DataFrame:
    total_pdf = len(df_pdf) if df_pdf is not None else 0
    total_excel = len(df_excel) if df_excel is not None else 0
    total_comp = len(df_comparacao) if df_comparacao is not None else 0
    colab_div = df_divergencias["Colaborador"].nunique() if not df_divergencias.empty else 0
    eventos_div = len(df_divergencias)
    ok = max(total_comp - colab_div, 0) if total_comp else 0
    return pd.DataFrame([
        {"Indicador": "Colaboradores identificados no PDF", "Quantidade": total_pdf},
        {"Indicador": "Colaboradores identificados na planilha", "Quantidade": total_excel},
        {"Indicador": "Colaboradores comparados", "Quantidade": total_comp},
        {"Indicador": "Colaboradores com divergência", "Quantidade": colab_div},
        {"Indicador": "Eventos divergentes", "Quantidade": eventos_div},
        {"Indicador": "Colaboradores sem divergência", "Quantidade": ok},
    ])


def gerar_excel(df_pdf: pd.DataFrame, df_excel: Optional[pd.DataFrame] = None, df_comparacao: Optional[pd.DataFrame] = None) -> bytes:
    """Gera um arquivo .XLSX real, nunca CSV.

    Importante: o retorno desta função precisa ser bytes de um workbook OOXML,
    que sempre começa com PK (arquivo ZIP interno do Excel). Isso evita o
    problema de abrir tudo em uma coluna só, que acontece quando o app entrega
    CSV/texto separado por vírgulas em vez de XLSX verdadeiro.
    """
    output = io.BytesIO()
    df_divergencias = montar_divergencias_limpas(df_comparacao)
    df_resumo = montar_resumo(df_pdf, df_excel, df_comparacao, df_divergencias)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Aba principal: só o que precisa de ação.
        df_divergencias.to_excel(writer, index=False, sheet_name="Divergências")
        df_resumo.to_excel(writer, index=False, sheet_name="Resumo")

        df_horas_falta = filtrar_horas_falta(df_divergencias)
        if not df_horas_falta.empty:
            df_horas_falta.to_excel(writer, index=False, sheet_name="Horas falta")

        if df_comparacao is not None and not df_comparacao.empty:
            df_comparacao.to_excel(writer, index=False, sheet_name="Completo")

        df_pdf.to_excel(writer, index=False, sheet_name="Eventos PDF")

        if df_excel is not None and not df_excel.empty:
            df_excel_export = df_excel.drop(columns=[c for c in df_excel.columns if c.endswith(" Min") or c == "Chave Nome"], errors="ignore")
            df_excel_export.to_excel(writer, index=False, sheet_name="Totais Ponto")

        aplicar_estilo_planilha(writer.book)

        # Deixa Divergências como primeira aba visível.
        if "Divergências" in writer.book.sheetnames:
            writer.book.active = writer.book.sheetnames.index("Divergências")

    dados = output.getvalue()
    # Validação de segurança: XLSX é um arquivo ZIP, portanto começa com b"PK".
    # Se não começar com PK, não envia o download como Excel para evitar arquivo quebrado/CSV.
    if not dados.startswith(b"PK"):
        raise ValueError("Falha ao gerar XLSX real. O arquivo gerado não está no formato Excel válido.")
    return dados

# -----------------------------
# Interface Streamlit
# -----------------------------

st.set_page_config(
    page_title="Auditoria de Folha | PDF x Ponto",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

PREMIUM_CSS = """
<style>
:root {
    --bg-main: #0f172a;
    --card-bg: rgba(255,255,255,0.92);
    --card-border: rgba(15, 23, 42, 0.08);
    --text-muted: #64748b;
    --brand: #1d4ed8;
    --brand-dark: #0f2f68;
    --danger: #dc2626;
    --warning: #d97706;
    --success: #16a34a;
}

/* Esconde barra nativa das tabelas do Streamlit; o download oficial é XLSX. */
[data-testid="stElementToolbar"] {display: none !important;}
button[title*="Download"] {display: none !important;}
button[aria-label*="Download"] {display: none !important;}

.block-container {
    padding-top: 1.4rem;
    padding-bottom: 3rem;
    max-width: 1450px;
}

section[data-testid="stSidebar"] > div {
    background: linear-gradient(180deg, #0f172a 0%, #111827 100%);
    color: white;
}
section[data-testid="stSidebar"] label, section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span, section[data-testid="stSidebar"] .stCaption {
    color: rgba(255,255,255,0.86) !important;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h1,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h2,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3 {
    color: white !important;
}

.premium-hero {
    position: relative;
    overflow: hidden;
    background:
        linear-gradient(135deg, rgba(15,23,42,.98) 0%, rgba(30,58,138,.96) 52%, rgba(15,23,42,.98) 100%);
    padding: 34px 38px 30px 38px;
    border-radius: 26px;
    color: white;
    box-shadow: 0 26px 70px rgba(15, 23, 42, 0.26);
    margin-bottom: 24px;
    border: 1px solid rgba(255,255,255,.14);
}
.premium-hero:before {
    content: "";
    position: absolute;
    width: 440px;
    height: 440px;
    right: -180px;
    top: -220px;
    background: radial-gradient(circle, rgba(96,165,250,.28) 0%, rgba(96,165,250,0) 65%);
}
.premium-hero:after {
    content: "";
    position: absolute;
    left: 38px;
    bottom: 0;
    width: 210px;
    height: 3px;
    background: linear-gradient(90deg, #60a5fa, #22c55e);
    border-radius: 99px 99px 0 0;
}
.premium-hero > * { position: relative; z-index: 1; }
.hero-kicker {
    color: #93c5fd;
    font-size: .78rem;
    font-weight: 850;
    letter-spacing: .16em;
    text-transform: uppercase;
    margin-bottom: 10px;
}
.premium-hero h1 {
    font-size: 2.35rem;
    line-height: 1.05;
    margin: 0 0 12px 0;
    font-weight: 900;
    letter-spacing: -0.055em;
    max-width: 980px;
}
.premium-hero p.hero-lead {
    margin: 0;
    max-width: 1010px;
    font-size: 1.02rem;
    line-height: 1.62;
    color: rgba(255,255,255,0.82);
}
.hero-summary {
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 14px;
    margin-top: 24px;
    max-width: 1120px;
}
.hero-summary div {
    background: rgba(255,255,255,.07);
    border: 1px solid rgba(255,255,255,.12);
    border-radius: 16px;
    padding: 14px 16px;
    backdrop-filter: blur(8px);
}
.hero-summary b {
    display: block;
    font-size: .82rem;
    color: #dbeafe;
    text-transform: uppercase;
    letter-spacing: .08em;
    margin-bottom: 6px;
}
.hero-summary span {
    display: block;
    color: rgba(255,255,255,.78);
    font-size: .88rem;
    line-height: 1.38;
}
@media (max-width: 900px) {
    .hero-summary { grid-template-columns: 1fr; }
    .premium-hero h1 { font-size: 1.85rem; }
}

.card {
    background: white;
    border: 1px solid var(--card-border);
    border-radius: 18px;
    padding: 18px 20px;
    box-shadow: 0 10px 30px rgba(15, 23, 42, .06);
}
.section-title {
    display:flex;
    align-items:center;
    gap:10px;
    font-weight: 800;
    font-size: 1.13rem;
    color:#0f172a;
    margin: 8px 0 12px 0;
}
.section-subtitle {
    color: var(--text-muted);
    font-size: .92rem;
    margin-top:-7px;
    margin-bottom: 14px;
}
.step-box {
    border: 1px dashed #cbd5e1;
    background: #f8fafc;
    border-radius: 16px;
    padding: 14px 16px;
    margin-bottom: 10px;
}
.step-box b {color:#0f172a;}
.step-box small {color:#64748b;}

.metric-card {
    background: white;
    border: 1px solid rgba(15, 23, 42, .08);
    border-radius: 18px;
    padding: 16px 18px;
    box-shadow: 0 8px 25px rgba(15, 23, 42, .06);
    min-height: 104px;
}
.metric-card .label {
    color: #64748b;
    font-size: .84rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .04em;
}
.metric-card .value {
    color: #0f172a;
    font-size: 1.85rem;
    font-weight: 850;
    letter-spacing: -0.04em;
    margin-top: 6px;
}
.metric-card .hint {
    color: #64748b;
    font-size: .82rem;
    margin-top: 1px;
}
.metric-danger {border-left: 6px solid #dc2626;}
.metric-success {border-left: 6px solid #16a34a;}
.metric-blue {border-left: 6px solid #2563eb;}
.metric-purple {border-left: 6px solid #7c3aed;}

.stButton > button, .stDownloadButton > button {
    border-radius: 14px !important;
    border: 0 !important;
    padding: 0.65rem 1.0rem !important;
    font-weight: 800 !important;
    box-shadow: 0 10px 24px rgba(29, 78, 216, .18);
}
.stDownloadButton > button {
    background: linear-gradient(135deg, #16a34a 0%, #15803d 100%) !important;
    color: white !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #2563eb 0%, #1e40af 100%) !important;
}

[data-testid="stFileUploader"] {
    border: 1px solid rgba(15, 23, 42, .08);
    border-radius: 18px;
    padding: 4px 12px 12px 12px;
    background: white;
    box-shadow: 0 10px 30px rgba(15, 23, 42, .05);
}

[data-testid="stDataFrameResizable"] {
    border-radius: 16px;
    overflow: hidden;
    border: 1px solid rgba(15, 23, 42, .08);
}

hr {margin: 1.2rem 0;}

.footer-note {
    color:#64748b;
    font-size:.84rem;
    text-align:center;
    padding: 18px 0 0 0;
}
</style>
"""
st.markdown(PREMIUM_CSS, unsafe_allow_html=True)

st.markdown(
    """
    <div class="premium-hero executive-hero">
        <div class="hero-kicker">AUDITORIA OPERACIONAL DE FOLHA</div>
        <h1>CONFERÊNCIA AUTOMÁTICA DE PONTO</h1>
        <p class="hero-lead">O sistema identifica os eventos de folha por colaborador, cruza com os totais da planilha de ponto e entrega resultado com foco nas divergências que exigem validação.</p>
        <div class="hero-summary">
            <div><b>Entrada</b><span>PDF da folha + planilha padrão de ponto</span></div>
            <div><b>Processamento</b><span>Leitura por colaborador, regras de auditoria e comparação automática</span></div>
            <div><b>Resultado</b><span>Relatório XLSX com divergências, resumo e auditoria completa</span></div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.markdown("### ⚙️ Configuração")
    perfil = st.selectbox(
        "Modelo de leitura",
        [
            "Automático - detectar todas as HEs do PDF",
            "Sucateamento - HE 75% / HE 100% / Noturno 27%",
            "Contrato - HE 50% / HE 100% / Noturno 27%",
            "Personalizado",
        ],
    )

    modo_automatico = perfil.startswith("Automático")

    if not modo_automatico:
        if perfil.startswith("Sucateamento"):
            he_1_percent = "75"
            he_2_percent = "100"
            noturno_percent = "27"
        elif perfil.startswith("Contrato - HE 50"):
            he_1_percent = "50"
            he_2_percent = "100"
            noturno_percent = "27"
        else:
            he_1_percent = st.text_input("Percentual da HE principal", value="75")
            he_2_percent = st.text_input("Percentual da HE secundária", value="100")
            noturno_percent = st.text_input("Percentual do adicional noturno", value="27")

        st.divider()
        st.caption("Colunas da planilha de ponto, considerando A=1, B=2, C=3...")
        col_he_1 = st.number_input("Coluna da HE principal", min_value=1, max_value=80, value=10, help="Padrão: J")
        col_he_2 = st.number_input("Coluna da HE secundária", min_value=1, max_value=80, value=11, help="Padrão: K")
        col_noturno = st.number_input("Coluna do adicional noturno", min_value=1, max_value=80, value=12, help="Padrão: L")
    else:
        st.info("Modo recomendado: o app detecta qualquer Hora Extra com percentual no PDF e cruza com o cabeçalho específico de cada colaborador na planilha.")
        comparar_noturno = st.checkbox("Comparar adicional noturno da planilha", value=True)

    st.divider()
    st.markdown("### 📌 Regras ativas")
    st.caption("• Dissídio antes de HE/Noturno é ignorado")
    st.caption("• Noturno XX% no PDF compara com Adicional Noturno no Excel")
    st.caption("• Coluna G >= 08:00 = falta em dia")
    st.caption("• Coluna G > 00:00 e < 08:00 = horas falta")

# Guarda os resultados no session_state para o download não apagar a análise.
for chave, valor_padrao in {
    "analise_concluida": False,
    "df_pdf": None,
    "df_excel": None,
    "df_comparacao": None,
    "df_divergencias": None,
    "excel_bytes": None,
    "mensagem_pdf": "",
    "mensagem_comparacao": "",
}.items():
    if chave not in st.session_state:
        st.session_state[chave] = valor_padrao

# Área de upload e processamento
st.markdown('<div class="section-title">📁 Arquivos para conferência</div>', unsafe_allow_html=True)
st.markdown('<div class="section-subtitle">Envie o PDF da folha e a planilha padrão de ponto. O sistema processa somente após clicar em “Processar conferência”.</div>', unsafe_allow_html=True)

col_upload_pdf, col_upload_excel = st.columns(2)
with col_upload_pdf:
    st.markdown('<div class="step-box"><b>1. PDF da folha</b><br><small>Arquivo da folha de pagamento com eventos por colaborador.</small></div>', unsafe_allow_html=True)
    arquivo_pdf = st.file_uploader("Selecione o PDF", type=["pdf"], label_visibility="collapsed")
with col_upload_excel:
    st.markdown('<div class="step-box"><b>2. Planilha de ponto</b><br><small>Modelo padrão com totais, HEs, adicional noturno e coluna G de débito.</small></div>', unsafe_allow_html=True)
    arquivo_excel = st.file_uploader("Selecione a planilha", type=["xlsx"], help="No modo automático, o app usa os cabeçalhos das colunas para identificar as HEs.", label_visibility="collapsed")

processar = False
if arquivo_pdf:
    col_btn, col_hint = st.columns([1, 3])
    with col_btn:
        processar = st.button("🚀 Processar conferência", type="primary", use_container_width=True)
    with col_hint:
        st.caption("☺ PROCESSO DE CONFERÊNCIA AUTOMÁTICO ☺.")
else:
    st.info("Envie pelo menos o PDF da folha para iniciar a leitura.")

if arquivo_pdf and processar:
    try:
        if modo_automatico:
            with st.spinner("Lendo PDF e detectando automaticamente todas as horas extras..."):
                df_pdf, percentuais_pdf = processar_pdf_dinamico(arquivo_pdf)
            msg_pdf = f"PDF processado. {len(df_pdf)} colaboradores identificados. HEs encontradas: {', '.join([p + '%' for p in percentuais_pdf]) if percentuais_pdf else 'nenhuma'}."

            df_excel = None
            df_comparacao = None
            percentuais_excel = []
            msg_comp = ""

            if arquivo_excel:
                with st.spinner("Lendo planilha e detectando colunas de horas extras pelo cabeçalho..."):
                    df_excel, percentuais_excel = processar_planilha_ponto_dinamica(arquivo_excel)
                percentuais_uniao = ordenar_percentuais(set(percentuais_pdf) | set(percentuais_excel))
                with st.spinner("Comparando PDF x Excel..."):
                    df_comparacao = comparar_dinamico(df_pdf, df_excel, percentuais_uniao, comparar_noturno=comparar_noturno)
                msg_comp = f"HEs consideradas na comparação: {', '.join([p + '%' for p in percentuais_uniao]) if percentuais_uniao else 'nenhuma'}."

        else:
            label_he_1 = f"Hora Extra {str(he_1_percent).replace('%', '').strip()}%"
            label_he_2 = f"Hora Extra {str(he_2_percent).replace('%', '').strip()}%"
            label_noturno = f"Noturno {str(noturno_percent).replace('%', '').strip()}%"
            eventos = montar_eventos_config(he_1_percent, he_2_percent, noturno_percent)
            colunas_pdf = montar_colunas_pdf(eventos)
            with st.spinner("Lendo PDF e identificando colaboradores/eventos..."):
                df_pdf = processar_pdf_manual(arquivo_pdf, eventos, colunas_pdf)
            msg_pdf = f"PDF processado. {len(df_pdf)} colaboradores identificados."

            df_excel = None
            df_comparacao = None
            msg_comp = ""
            if arquivo_excel:
                with st.spinner("Lendo planilha de ponto e comparando com o PDF..."):
                    df_excel = processar_planilha_ponto_manual(arquivo_excel, label_he_1, label_he_2, label_noturno, int(col_he_1), int(col_he_2), int(col_noturno))
                    df_comparacao = comparar_manual(df_pdf, df_excel, label_he_1, label_he_2, label_noturno)

        if df_pdf.empty:
            st.session_state["analise_concluida"] = False
            st.warning("Nenhum colaborador foi identificado no PDF. Verifique se o PDF possui texto extraível.")
        else:
            df_divergencias = montar_divergencias_limpas(df_comparacao) if df_comparacao is not None else pd.DataFrame()
            excel_bytes = gerar_excel(df_pdf, df_excel, df_comparacao)

            st.session_state["df_pdf"] = df_pdf
            st.session_state["df_excel"] = df_excel
            st.session_state["df_comparacao"] = df_comparacao
            st.session_state["df_divergencias"] = df_divergencias
            st.session_state["excel_bytes"] = excel_bytes
            st.session_state["mensagem_pdf"] = msg_pdf
            st.session_state["mensagem_comparacao"] = msg_comp
            st.session_state["analise_concluida"] = True
            st.toast("Conferência concluída com sucesso.", icon="✅")

    except Exception as e:
        st.error(f"Erro ao processar os arquivos: {e}")

# Renderiza os resultados fora do botão para não sumirem no rerun do download.
if st.session_state.get("analise_concluida"):
    df_pdf = st.session_state["df_pdf"]
    df_excel = st.session_state["df_excel"]
    df_comparacao = st.session_state["df_comparacao"]
    df_divergencias_view = st.session_state["df_divergencias"]
    excel_bytes = st.session_state["excel_bytes"]

    if st.session_state.get("mensagem_pdf"):
        st.success(st.session_state["mensagem_pdf"])
    if st.session_state.get("mensagem_comparacao"):
        st.info(st.session_state["mensagem_comparacao"])

    total_pdf = len(df_pdf) if df_pdf is not None else 0
    total_excel = len(df_excel) if df_excel is not None else 0
    total_div = len(df_divergencias_view) if df_divergencias_view is not None else 0
    colab_div = df_divergencias_view["Colaborador"].nunique() if df_divergencias_view is not None and not df_divergencias_view.empty and "Colaborador" in df_divergencias_view.columns else 0
    total_comp = len(df_comparacao) if df_comparacao is not None else 0
    total_ok = max(total_comp - colab_div, 0) if total_comp else 0

    st.markdown('<div class="section-title">📊 Resultado da conferência</div>', unsafe_allow_html=True)
    m1, m2, m3, m4, m5 = st.columns(5)
    with m1:
        st.markdown(f'<div class="metric-card metric-blue"><div class="label">PDF</div><div class="value">{total_pdf}</div><div class="hint">colaboradores lidos</div></div>', unsafe_allow_html=True)
    with m2:
        st.markdown(f'<div class="metric-card metric-blue"><div class="label">Planilha</div><div class="value">{total_excel}</div><div class="hint">colaboradores lidos</div></div>', unsafe_allow_html=True)
    with m3:
        st.markdown(f'<div class="metric-card metric-danger"><div class="label">Divergências</div><div class="value">{total_div}</div><div class="hint">eventos para ação</div></div>', unsafe_allow_html=True)
    with m4:
        st.markdown(f'<div class="metric-card metric-purple"><div class="label">Colab. divergentes</div><div class="value">{colab_div}</div><div class="hint">pessoas com ajuste</div></div>', unsafe_allow_html=True)
    with m5:
        st.markdown(f'<div class="metric-card metric-success"><div class="label">Sem divergência</div><div class="value">{total_ok}</div><div class="hint">comparados OK</div></div>', unsafe_allow_html=True)

    st.divider()

    if df_comparacao is not None and not df_comparacao.empty:
        if df_divergencias_view is None or df_divergencias_view.empty:
            st.success("Nenhuma divergência encontrada. Todos os eventos comparados estão OK.")
        else:
            st.markdown('<div class="section-title">🚨 Divergências encontradas</div>', unsafe_allow_html=True)
            st.markdown('<div class="section-subtitle">Visão principal: apenas ocorrências que precisam ser avaliadas pela equipe.</div>', unsafe_allow_html=True)

            busca = st.text_input("Buscar colaborador ou evento", placeholder="Digite parte do nome, evento ou status...")
            df_div_filtrado = df_divergencias_view.copy()
            if busca:
                busca_norm = normalizar_texto(busca)
                mask = df_div_filtrado.apply(lambda row: busca_norm in normalizar_texto(" ".join(map(str, row.values))), axis=1)
                df_div_filtrado = df_div_filtrado[mask]

            df_horas_falta_view = filtrar_horas_falta(df_div_filtrado)
            if not df_horas_falta_view.empty:
                st.warning(f"Horas falta: {len(df_horas_falta_view)} divergência(s) encontrada(s) pela regra da coluna G.")

            st.dataframe(df_div_filtrado, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhuma comparação gerada. O XLSX terá os eventos encontrados no PDF.")

    st.download_button(
        label="⬇️ Baixar planilha OFICIAL em XLSX",
        data=excel_bytes,
        file_name="conferencia_pdf_x_ponto.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Use este botão para baixar o arquivo Excel verdadeiro (.xlsx). A análise ficará salva na tela após o download.",
        use_container_width=True,
    )

    tab_div, tab_pdf, tab_ponto, tab_completo = st.tabs(["🚨 Divergências", "📄 Eventos PDF", "📘 Totais Ponto", "🔎 Auditoria completa"])

    with tab_div:
        if df_divergencias_view is not None and not df_divergencias_view.empty:
            st.dataframe(df_divergencias_view, use_container_width=True, hide_index=True)
        else:
            st.success("Sem divergências para exibir.")

    with tab_pdf:
        colunas_eventos = [c for c in df_pdf.columns if c not in ["Código", "Colaborador", "Página"]]
        df_pdf_com_evento = df_pdf[df_pdf[colunas_eventos].apply(lambda row: any(str(v).strip() for v in row), axis=1)] if colunas_eventos else df_pdf
        st.dataframe(df_pdf_com_evento if not df_pdf_com_evento.empty else df_pdf, use_container_width=True, hide_index=True)

    with tab_ponto:
        if df_excel is not None:
            if df_excel.empty:
                st.warning("Nenhum total foi identificado na planilha. Verifique se existe linha 'Colaborador', cabeçalho com HEs e linha 'TOTAIS'.")
            else:
                df_excel_view = df_excel.drop(columns=[c for c in df_excel.columns if c.endswith(" Min") or c == "Chave Nome"], errors="ignore")
                st.dataframe(df_excel_view, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhuma planilha de ponto foi enviada.")

    with tab_completo:
        if df_comparacao is not None and not df_comparacao.empty:
            st.dataframe(df_comparacao, use_container_width=True, hide_index=True)
        else:
            st.info("A comparação completa aparece aqui quando PDF e planilha são processados juntos.")

st.markdown('<div class="footer-note">Auditoria operacional • PDF x ponto • XLSX oficial • Jhonnathan Lamberti Pereira</div>', unsafe_allow_html=True)
